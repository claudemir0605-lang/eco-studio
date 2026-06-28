# -*- coding: utf-8 -*-
"""
Gera a planilha pessoal de controle de gastos e consumo mensal.
Execute:  python3 gerar_planilha.py
Saida:    Controle_de_Gastos.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ANO = 2026  # ano de referencia do resumo mensal
ARQUIVO = "Controle_de_Gastos.xlsx"

# ---------------------------------------------------------------- estilos
AZUL = "1F4E78"
AZUL_CLARO = "DDEBF7"
CINZA = "F2F2F2"
VERDE = "548235"
VERMELHO = "C00000"

titulo = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
cabec = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
neg = Font(name="Calibri", size=11, bold=True)
normal = Font(name="Calibri", size=11)

fill_titulo = PatternFill("solid", fgColor=AZUL)
fill_cabec = PatternFill("solid", fgColor=AZUL)
fill_total = PatternFill("solid", fgColor=AZUL_CLARO)
fill_zebra = PatternFill("solid", fgColor=CINZA)

centro = Alignment(horizontal="center", vertical="center")
esq = Alignment(horizontal="left", vertical="center")

fino = Side(style="thin", color="BFBFBF")
borda = Border(left=fino, right=fino, top=fino, bottom=fino)

MOEDA = 'R$ #,##0.00'

wb = Workbook()

# ================================================================ CATEGORIAS
ws_cat = wb.active
ws_cat.title = "Categorias"

categorias = [
    "Moradia (aluguel/condomínio)",
    "Contas (água/luz/gás/internet)",
    "Alimentação (mercado)",
    "Refeições fora",
    "Transporte (combustível/ônibus)",
    "Saúde (farmácia/consultas)",
    "Educação",
    "Lazer",
    "Vestuário",
    "Assinaturas/Serviços",
    "Cuidados pessoais",
    "Pets",
    "Imprevistos",
    "Outros",
]
pagamentos = ["Dinheiro", "Pix", "Débito", "Crédito", "Boleto", "Vale"]

ws_cat["A1"] = "Categorias"
ws_cat["C1"] = "Formas de pagamento"
for c in ("A1", "C1"):
    ws_cat[c].font = cabec
    ws_cat[c].fill = fill_cabec
    ws_cat[c].alignment = centro
for i, cat in enumerate(categorias, start=2):
    ws_cat.cell(row=i, column=1, value=cat).font = normal
for i, pg in enumerate(pagamentos, start=2):
    ws_cat.cell(row=i, column=3, value=pg).font = normal
ws_cat.column_dimensions["A"].width = 34
ws_cat.column_dimensions["B"].width = 3
ws_cat.column_dimensions["C"].width = 22
ws_cat.sheet_view.showGridLines = False

n_cat = len(categorias)
n_pag = len(pagamentos)
ref_cat = f"Categorias!$A$2:$A${n_cat+1}"
ref_pag = f"Categorias!$C$2:$C${n_pag+1}"

# ================================================================ LANÇAMENTOS
ws = wb.create_sheet("Lançamentos")
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:E1")
ws["A1"] = "Lançamentos diários — registre aqui cada gasto"
ws["A1"].font = titulo
ws["A1"].fill = fill_titulo
ws["A1"].alignment = centro
ws.row_dimensions[1].height = 26

cabecalhos = ["Data", "Categoria", "Descrição", "Forma de pagamento", "Valor (R$)"]
for col, txt in enumerate(cabecalhos, start=1):
    c = ws.cell(row=2, column=col, value=txt)
    c.font = cabec
    c.fill = fill_cabec
    c.alignment = centro
    c.border = borda

larguras = {"A": 14, "B": 34, "C": 38, "D": 20, "E": 14}
for col, w in larguras.items():
    ws.column_dimensions[col].width = w

PRIMEIRA = 3
ULTIMA = 502  # ~500 linhas para lançar o ano todo
for r in range(PRIMEIRA, ULTIMA + 1):
    for col in range(1, 6):
        cell = ws.cell(row=r, column=col)
        cell.border = borda
        cell.font = normal
        if r % 2 == 0:
            cell.fill = fill_zebra
    ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
    ws.cell(row=r, column=5).number_format = MOEDA

# validações (listas suspensas)
dv_cat = DataValidation(type="list", formula1=f"={ref_cat}", allow_blank=True)
dv_pag = DataValidation(type="list", formula1=f"={ref_pag}", allow_blank=True)
ws.add_data_validation(dv_cat)
ws.add_data_validation(dv_pag)
dv_cat.add(f"B{PRIMEIRA}:B{ULTIMA}")
dv_pag.add(f"D{PRIMEIRA}:D{ULTIMA}")

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:E{ULTIMA}"

# ================================================================ RESUMO MENSAL
rs = wb.create_sheet("Resumo Mensal")
rs.sheet_view.showGridLines = False

rs.merge_cells("A1:O1")
rs["A1"] = f"Resumo mensal por categoria — {ANO}"
rs["A1"].font = titulo
rs["A1"].fill = fill_titulo
rs["A1"].alignment = centro
rs.row_dimensions[1].height = 26

rs["A2"] = "Categoria"
rs["A2"].font = cabec
rs["A2"].fill = fill_cabec
rs["A2"].alignment = centro
rs["A2"].border = borda

# colunas dos 12 meses (B..M) com datas do dia 1 de cada mês
for m in range(1, 13):
    col = 1 + m  # B = mês 1
    cell = rs.cell(row=2, column=col, value=f"=DATE({ANO},{m},1)")
    cell.number_format = "mmm/yy"
    cell.font = cabec
    cell.fill = fill_cabec
    cell.alignment = centro
    cell.border = borda
    rs.column_dimensions[get_column_letter(col)].width = 11

# coluna Total (N) e Média (O)
total_cl = get_column_letter(14)
media_cl = get_column_letter(15)
for col, txt in ((14, "Total"), (15, "Média/mês")):
    c = rs.cell(row=2, column=col, value=txt)
    c.font = cabec
    c.fill = fill_cabec
    c.alignment = centro
    c.border = borda
rs.column_dimensions["A"].width = 34
rs.column_dimensions[total_cl].width = 13
rs.column_dimensions[media_cl].width = 12

faixa_data = "Lançamentos!$A:$A"
faixa_cat = "Lançamentos!$B:$B"
faixa_val = "Lançamentos!$E:$E"

linha0 = 3
for i, cat in enumerate(categorias):
    r = linha0 + i
    a = rs.cell(row=r, column=1, value=cat)
    a.font = normal
    a.border = borda
    a.alignment = esq
    for m in range(1, 13):
        col = 1 + m
        cl = get_column_letter(col)
        formula = (
            f"=SUMIFS({faixa_val},{faixa_cat},$A{r},"
            f"{faixa_data},\">=\"&{cl}$2,{faixa_data},\"<\"&EDATE({cl}$2,1))"
        )
        c = rs.cell(row=r, column=col, value=formula)
        c.number_format = MOEDA
        c.font = normal
        c.border = borda
    # total da categoria no ano
    c = rs.cell(row=r, column=14, value=f"=SUM(B{r}:M{r})")
    c.number_format = MOEDA
    c.font = neg
    c.border = borda
    c.fill = fill_total
    # média mensal (só meses com gasto)
    c = rs.cell(row=r, column=15,
                value=f"=IFERROR(AVERAGEIF(B{r}:M{r},\">0\"),0)")
    c.number_format = MOEDA
    c.font = normal
    c.border = borda

# linha de totais
rtot = linha0 + len(categorias)
ct = rs.cell(row=rtot, column=1, value="TOTAL DO MÊS")
ct.font = Font(bold=True, color="FFFFFF")
ct.fill = fill_cabec
ct.alignment = esq
ct.border = borda
for col in range(2, 16):
    cl = get_column_letter(col)
    c = rs.cell(row=rtot, column=col,
                value=f"=SUM({cl}{linha0}:{cl}{rtot-1})")
    c.number_format = MOEDA
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill_cabec
    c.border = borda

rs.freeze_panes = "B3"

# ================================================================ PAINEL
pn = wb.create_sheet("Painel", 0)  # primeira aba
pn.sheet_view.showGridLines = False
pn.merge_cells("A1:D1")
pn["A1"] = "💰 Controle de Gastos Pessoal"
pn["A1"].font = titulo
pn["A1"].fill = fill_titulo
pn["A1"].alignment = centro
pn.row_dimensions[1].height = 30

cards = [
    ("Total gasto no ano", f"=SUM(Lançamentos!E:E)", MOEDA),
    ("Nº de lançamentos", f"=COUNT(Lançamentos!E:E)", "0"),
    ("Maior gasto", f"=IFERROR(MAX(Lançamentos!E:E),0)", MOEDA),
    ("Gasto médio por lançamento",
     f"=IFERROR(AVERAGE(Lançamentos!E:E),0)", MOEDA),
    ("Gasto do mês atual",
     "=SUMIFS(Lançamentos!E:E,Lançamentos!A:A,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),"
     "Lançamentos!A:A,\"<\"&EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),1))",
     MOEDA),
]
linha = 3
for rotulo, formula, fmt in cards:
    pn.cell(row=linha, column=1, value=rotulo).font = neg
    v = pn.cell(row=linha, column=3, value=formula)
    v.number_format = fmt
    v.font = Font(bold=True, size=12, color=AZUL)
    pn.cell(row=linha, column=1).fill = fill_total
    pn.cell(row=linha, column=3).fill = fill_total
    linha += 1

pn.column_dimensions["A"].width = 30
pn.column_dimensions["B"].width = 3
pn.column_dimensions["C"].width = 18

pn.cell(row=linha + 1, column=1,
        value="Como usar:").font = neg
instrucoes = [
    "1. Vá na aba \"Lançamentos\" e registre cada gasto: data, categoria, descrição, forma e valor.",
    "2. As categorias e formas de pagamento aparecem em listas suspensas (clique na célula).",
    "3. A aba \"Resumo Mensal\" soma tudo automaticamente por categoria e por mês.",
    "4. Este \"Painel\" mostra os números gerais. Edite categorias na aba \"Categorias\".",
]
for i, txt in enumerate(instrucoes):
    pn.cell(row=linha + 2 + i, column=1, value=txt).font = normal
    pn.merge_cells(start_row=linha + 2 + i, start_column=1,
                   end_row=linha + 2 + i, end_column=4)

wb.save(ARQUIVO)
print("Planilha gerada:", ARQUIVO)
